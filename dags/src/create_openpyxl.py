import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl import Workbook

BASE_DIR = Path(__file__).parent.parent.parent

EXCEL_FILE = BASE_DIR/'data'/'online_retail_II.xlsx'
OUTPUT_DIR = BASE_DIR/'output'
OUTPUT_DIR.mkdir(parents=True,exist_ok=True)

def create_files_from_sheets(input_path=EXCEL_FILE, output_dir=OUTPUT_DIR):
    # Ensure output directory exists
    os.makedirs(output_dir, exist_ok=True)
    
    # Load the input workbook
    wb = load_workbook(input_path)
    
    # Iterate over each sheet in the workbook
    for sheet in wb.sheetnames:
        # Create a new workbook
        wb_new = Workbook()
        
        # Copy the current sheet to the new workbook
        ws_old = wb[sheet]
        ws_new = wb_new.active
        ws_new.title = sheet
        for row in ws_old.iter_rows():
            for cell in row:
                ws_new[cell.coordinate].value = cell.value
        
        # Save the new workbook with the name of the current sheet
        output_path = os.path.join(output_dir, f'{sheet}.xlsx')
        wb_new.save(output_path)
        
        # Close the new workbook
        wb_new.close()

    # Close the input workbook
    wb.close()

# Example usage
if __name__ == "__main__":
    unzip_f = create_files_from_sheets(EXCEL_FILE, OUTPUT_DIR)